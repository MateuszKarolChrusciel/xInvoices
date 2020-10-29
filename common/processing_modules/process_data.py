import openpyxl
import os

from common import decorators as decors
from common import functions as funcs
from common import program_objs as objs
from common.functions import multiprocess as mp_func
from common.processing_modules import process_files


# https://en.wikipedia.org/wiki/Pigeonhole_sort
def pigeonhole_sort(lst) -> None:
    """Sort list of (key, value) pairs by key."""
    base = min(key for key, value in lst)
    size = max(key for key, value in lst) - base + 1

    pigeonholes = [[] for _ in range(size)]

    for key, value in lst:
        i = key - base
        pigeonhole = pigeonholes[i]
        pigeonhole.append([key, value])

    i = 0
    for pigeonhole in pigeonholes:
        for element in pigeonhole:
            lst[i] = element
            i += 1


def extract_cells(cells):
    text = ""
    for cell in cells:
        text += cell.value


@decors.timeit
def process(pdfs, xlsxs):
    if not pdfs or not xlsxs:
        raise FileNotFoundError

    if not os.name == 'nt':
        raise EnvironmentError

    wbs = [openpyxl.load_workbook(xlsx) for xlsx in xlsxs]
    # wbs = mp_func(process_files.read_excel, xlsxs)

    print(wbs[0][0])

    lst_of_lst_of_cells = list()
    for counter, wb in enumerate(wbs):
        ws = wb[1]
        lst_of_lst_of_cells[counter] = [ws.cell(row, column) for row in range(ws.max_row) for column in
                                        range(ws.max_column)]

    old_invoices_obj = mp_func(objs.Invoice, extract_cells([cells for cells in lst_of_lst_of_cells]))

    new_invoices_obj = mp_func(objs.Invoice, mp_func(process_files.read_pdf, pdfs))

    print(len(old_invoices_obj))
    print(len(new_invoices_obj))

    pigeonhole_sort(old_invoices_obj)
    pigeonhole_sort(new_invoices_obj)
