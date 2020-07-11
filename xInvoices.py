import datetime
import openpyxl
import os
import pdfminerextractor as pex
import re
import shutil
import sys
import time
from openpyxl import utils as utl

try:
    import NIPy
except ImportError:
    NIPy = None

CURRENT_YEAR = datetime.datetime.now().strftime("%Y")

BASE_PATH = os.path.dirname(os.path.realpath(__file__))

DATA_PATH = f"{BASE_PATH}/DATA/"
OUTPUT_PATH = f"{BASE_PATH}/OUTPUT/"

DEFAULT_EXCEL_FILE_NAME = f"Kupno-sprzedaż {CURRENT_YEAR}.xlsx"


class Invoice:
    def __init__(self, invoice_data):
        self._raw_content = invoice_data

        self._true_raw_content = self.get_true_raw_content()

        self._buyer_nip = self.get_buyer_nip()
        self._invoice_id = self.get_invoice_id()
        self._payment_date = self.get_payment_date()
        self._payment_method = self.get_payment_method()
        self._raw_prices = self._get_raw_prices()
        self._signing_date = self.get_signing_date()

    @property
    def correct(self):
        if self.invoice_id[0] == "K":
            return True
        else:
            return False

    @property
    def raw_content(self):
        return self._raw_content

    @property
    def true_raw_content(self):
        return self._true_raw_content

    @property
    def buyer_nip(self):
        return self._buyer_nip

    @property
    def gross_value(self):
        try:
            return self.raw_prices["gross"]
        except KeyError:
            return self.net_value + self.vat_value

    @property
    def invoice_id(self):
        return self._invoice_id

    @property
    def net_value(self):
        return self.raw_prices["net"]

    @property
    def payment_method(self):
        return self._payment_method

    @property
    def payment_date(self):
        return self._payment_date

    @property
    def raw_prices(self):
        return self._raw_prices

    @property
    def signing_date(self):
        return self._signing_date

    @property
    def vat_value(self):
        return self.raw_prices["vat"]

    def get_buyer_nip(self):
        pattern1 = r"NIP\W*\d{10}"
        pattern2 = r"NIP\W*\d{3}\W\d{3}\W\d{2}\W\d{2}"
        pattern3 = r"NIP\W*\d{3}\W\d{2}\W\d{3}\W\d{2}"
        pattern4 = r"NIP\W*\d{3}\W\d{2}\W\d{2}\W\d{3}"
        pattern5 = r"NIP\W*\d{2}\W\d{2}\W\d{3}\W\d{3}"
        pattern6 = r"NIP\W*\d{2}\W\d{3}\W\d{2}\W\d{3}"
        pattern7 = r"NIP\W*\d{2}\W\d{3}\W\d{3}\W\d{2}"

        patterns = [pattern1, pattern2, pattern3, pattern4, pattern5, pattern6, pattern7]

        results = []

        for pattern in patterns:
            results.append(re.findall(pattern, self.raw_content))

        result = None

        for rl in results:
            if rl:
                for rs in rl:
                    if "773-156-23-77" in rs or "7731562377" in rs:
                        pass
                    else:
                        result = rs
            else:
                pass

        correct_result = ""

        try:
            for sign in result:
                if not sign.isdigit():
                    pass
                else:
                    correct_result = f"{correct_result}{sign}"
        except TypeError:
            raise NotImplementedError

        return f"{correct_result}"

    def get_true_raw_content(self):
        split_content = self.raw_content.split("\n")

        for counter, line in enumerate(split_content):
            if not line:
                split_content.pop(counter)
            else:
                while "\n" in line:
                    line -= "\n"

        return split_content

    def get_invoice_id(self):
        pattern = r"(?:numer)(?::*)(?:\s*)(\w*(/)*\w*((/)*\w*)*)"

        results = re.findall(pattern, self.raw_content)

        if len(results) == 1:
            result = results[0][0]
        else:
            raise NotImplementedError

        return result

    def get_payment_date(self):
        def find_line_with_pd(lines):
            line_w_pm = None

            for line in lines:
                if "termin płatności" in line.lower():
                    line_w_pm = line
                    break
                else:
                    pass

            return line_w_pm

        def get_raw_pd(line_w_pd):
            indexes = []

            for counter, sign in enumerate(line_w_pd):
                if sign.isdigit():
                    indexes.append(counter)

            pd_raw_str = line_w_pd[indexes[0]: None]

            try:
                y = int(pd_raw_str[0:4])
                m = int(pd_raw_str[5:7])
            except ValueError:
                raise NotImplementedError
            try:
                d = int(pd_raw_str[8:10])
            except ValueError or KeyError:
                d = int(pd_raw_str[8:None])

            return y, m, d

        line_with_payment_date = find_line_with_pd(self.true_raw_content)

        try:
            year, month, day = get_raw_pd(line_with_payment_date)
        except NotImplementedError:
            return None

        payment_date = datetime.date(year=year, month=month, day=day)

        return payment_date

    def get_payment_method(self):
        def find_line_with_pm(lines):
            line_w_pm = None

            for line in lines:
                if "płatność" in line.lower():
                    line_w_pm = line
                    break
                else:
                    pass

            if line_w_pm is None:
                return "nic"
            else:
                return line_w_pm

        line_with_pm = find_line_with_pm(self.true_raw_content)

        if "gotówka" in line_with_pm.lower():
            pm = "Gotówka"
        elif "przelew" in line_with_pm.lower():
            pm = "Przelew"
        elif "nic" in line_with_pm.lower():
            pm = ""
        else:
            raise ValueError

        return pm.upper()

    def get_signing_date(self):
        def find_line_w_sd(lines):
            line_w_sd = None
            found_line_w_sd = False

            while True:
                for line in lines:
                    if "wystawienia" in line.lower():
                        line_w_sd = line
                        found_line_w_sd = True
                        break
                if found_line_w_sd:
                    break
                else:
                    raise ValueError

            return line_w_sd

        def get_raw_sd(line_w_sd):
            indexes = []

            for counter, sign in enumerate(line_w_sd):
                if sign.isdigit():
                    indexes.append(counter)

            sd_raw_str = line_w_sd[indexes[0]: None]

            y = int(sd_raw_str[0:4])
            m = int(sd_raw_str[5:7])
            try:
                d = int(sd_raw_str[8:10])
            except ValueError or KeyError:
                d = int(sd_raw_str[8:None])

            return y, m, d

        line_with_signing_date = find_line_w_sd(self.true_raw_content)

        year, month, day = get_raw_sd(line_with_signing_date)

        date_when_signed = datetime.date(year=year, month=month, day=day)

        return date_when_signed

    def _get_raw_prices(self):
        regex = r"(\d{1,100}\s{1,100})*\d*[,]\d\d\s*PLN"

        matches = re.finditer(regex, self.raw_content, re.MULTILINE)

        finds = []

        for match in matches:
            finds.append(match.group())

        real_finds = []

        for find in finds:
            if find in real_finds:
                pass
            else:
                real_finds.append(find)

        for counter, find in enumerate(real_finds):
            if find.endswith(" PLN"):
                real_finds[counter] = find[:-4]

        prics = list()

        for find in real_finds:
            prics.append(float(find.replace(",", ".").replace(" ", "").replace("\n", "")))

        prices = dict()

        prices["gross"] = max(prics)
        prices["vat"] = min(prics)
        prices["net"] = prices["gross"] - prices["vat"]

        return prices


def _ask_user_if_one_wants_to_continue():
    print("")
    while True:
        user_input = input("Czy chcesz kontynuować? (T/N): ")

        if user_input.upper() == "T":
            break
        elif user_input.upper() == "N":
            leave()
        else:
            print("\nWpisano złą wartość. Spróbuj ponownie.")

    clear_console()


def convert_pdf_data_files_to_invoice_obj_list(pdf_files):
    print("Konwertowanie plików pdf do tekstu...")

    invoices = []
    files_to_remove = []

    num_of_files = len(pdf_files)
    num_of_files_converted = num_of_files

    for file in pdf_files:
        clear_console()

        print(f"Konwertowanie pliku \"{file}\"...")

        file_path = pdf_files[file]

        try:
            invoices.append(Invoice(pex.convert_pdf_to_txt(file_path)))
        except (KeyError, NotImplementedError):
            files_to_remove.append(file)
            num_of_files_converted -= 1
            print(f"Nie udało się przekonwertować! Pomijam plik...")
        else:
            print("\nPlik przekonwertowano pomyślnie...")

    for file in files_to_remove:
        pdf_files.pop(file)

    clear_console()

    print(f"\nPrzekonwertowanych plików: {num_of_files_converted}/{num_of_files}")

    return invoices


def enter_data_to_workbook(invoices):
    for invoice in invoices:
        clear_console()

        print(f"Wpisuję dane faktury {invoice.invoice_id} do pliku wyjściowego...")

        last_cell_with_data_in_rows = 0

        for cll in worksheet["1"]:
            column_letter = utl.get_column_letter(cll.column)
            for counter, cell in enumerate(worksheet[f"{column_letter}"], 1):
                if not cell.value:
                    pass
                else:
                    if last_cell_with_data_in_rows < counter:
                        last_cell_with_data_in_rows = counter
                    else:
                        pass

        cells_in_c_with_data = {}
        for counter, cell in enumerate(worksheet["C"][:last_cell_with_data_in_rows]):
            cells_in_c_with_data[counter] = cell

        try:
            for cell_id in reversed(cells_in_c_with_data):
                cell = cells_in_c_with_data[cell_id]
                if cell.value == invoice.invoice_id:
                    raise ValueError
                else:
                    pass
        except ValueError:
            print(f"\nDane faktury o numerze {invoice.invoice_id} znajdują się już w folderze wyjściowym! Pomijam...")
            continue

        try:
            for cell_id in reversed(cells_in_c_with_data):
                for cl, cll in zip(worksheet[f"{cell_id + 1}"], worksheet[f"{cell_id + 2}"]):
                    if not invoice.correct:
                        if type(cl.value) == datetime.datetime:
                            cll.value = cl.value.date()
                            cl.value = cl.value.date()
                        else:
                            cll.value = cl.value

                        # print(invoice.payment_date)
                        if type(worksheet[f"A{cell_id + 1}"].value) == datetime.date \
                                and type(invoice.payment_date) == datetime.date:
                            if invoice.payment_date > worksheet[f"A{cell_id + 1}"].value:
                                correct_row = cell_id + 2
                                worksheet[f"A{correct_row}"] = invoice.signing_date

                                if NIPy:
                                    try:
                                        worksheet[f"B{correct_row}"] = NIPy.nipy[invoice.buyer_nip]
                                    except KeyError:
                                        worksheet[f"B{correct_row}"] = invoice.buyer_nip
                                else:
                                    worksheet[f"B{correct_row}"] = invoice.buyer_nip

                                worksheet[f"C{correct_row}"] = invoice.invoice_id
                                worksheet[f"D{correct_row}"] = invoice.net_value
                                worksheet[f"E{correct_row}"] = invoice.vat_value
                                worksheet[f"G{correct_row}"] = invoice.payment_method
                                worksheet[f"H{correct_row}"] = invoice.payment_date
                                raise StopIteration
                            else:
                                pass
                        else:
                            pass
        except StopIteration:
            print(f"\nPomyślnie wpisano dane faktury {invoice.invoice_id} do pliku wyjściowego...")
            continue

        print(f"\nNie udało się wpisać danych faktury {invoice.invoice_id} do pliku wyjściowego...")


def enter_gross_value_cells_formulas():
    for counter, cell in enumerate(worksheet["F"], 1):
        try:
            float(worksheet[f"D{counter}"].value)
            float(worksheet[f"E{counter}"].value)
        except TypeError:
            pass
        except ValueError:
            pass
        else:
            cell.value = f"=D{counter}+E{counter}"
            continue

        if counter != 1:
            cell.value = None


def format_output_file():
    enter_gross_value_cells_formulas()


def get_pdf_data_files_name_n_path():
    result = {}
    pdf_data_files_names = []
    pdf_data_files_paths = []

    for file in os.listdir(DATA_PATH):
        full_file_path = f"{DATA_PATH}/{file}"
        if file.endswith(".pdf"):
            pdf_data_files_paths.append(full_file_path)
            pdf_data_files_names.append(file)
        else:
            pass

    for name, path in zip(pdf_data_files_names, pdf_data_files_paths):
        result[name] = path

    return result


def inform_user_about_overwriting():
    print("Ten program może zmodyfikować, nadpisać lub usunąć pliki w folderze:")
    print(f"\n{BASE_PATH}\n")
    print("oraz jego podfolderach!")
    _ask_user_if_one_wants_to_continue()


def process_data():
    enter_data_to_workbook(convert_pdf_data_files_to_invoice_obj_list(get_pdf_data_files_name_n_path()))

    clear_console()


def save_excel_output_file():
    print("Zapisuję ostateczną wersję pliku wyjściowego...")

    workbook.save(f"{OUTPUT_PATH}/{excel}")

    print("Zapisano plik wyjściowy...")

    clear_console()


def verify_pdf_files_existence():
    print("Wyszukiwanie plików PDF w foledrze danych...")

    pdf_file_found = False
    while True:
        for file in os.listdir(DATA_PATH):
            if file.endswith(".pdf"):
                pdf_file_found = True
                break

        if pdf_file_found:
            break
        else:
            pass

        print("Nie znaleziono żadnych plików PDF! Umieść plik(i) PDF w folderze danych!")
        input("Wciśnij \"Enter\" aby zamknąć program: ")

        leave()

    print("Znaleziono pliki PDF...")

    clear_console()


def verify_data_path():
    print("Weryfikacja stanu folderu danych...")

    is_data_path = bool(os.path.isdir(DATA_PATH))

    if not is_data_path:
        print("Folder danych nie znaleziony!")
        print("Tworzenie folderu danych...")

        os.makedirs(DATA_PATH)
        is_data_dir_empty = True

        print("Utworzono folder danych...")
    else:
        is_data_dir_not_empty = bool(len(os.listdir(DATA_PATH)))

        if not is_data_dir_not_empty:
            is_data_dir_empty = True
        else:
            is_data_dir_empty = False

    if is_data_dir_empty:
        clear_console()
        print("Wystąpił błąd! Nie znaleziono plików w folderze danych!")
        input("\nWciśnij \"Enter\" aby zamknąć program: ")

        leave()

    else:
        print("Pomyślnie zweryfikowano folder danych...")

    clear_console()


def verify_environment():
    verify_data_path()
    verify_output_path()
    verify_pdf_files_existence()
    excl = verify_excel_input_file()

    return excl


def verify_excel_input_file():
    def verify_default_excel_file_name_correctness():
        chosen_excel_file_name = DEFAULT_EXCEL_FILE_NAME
        print(f"\nAktualna nazwa docelowego pliku arkusza wejsciowego: \"{chosen_excel_file_name}\"")

        while True:
            user_decision = input("Czy zmienić nazwę docelowego pliku arkusza wejsciowego? (T/N): ")
            clear_console()
            if user_decision.upper() == "T":
                entered_name = input("\nWprowadź nową nazwę dla docelowego pliku arkusza wejsciowego: ")
                clear_console()
                print(f"\nNowa nazwa dla docelowego pliku arkusza wejsciowego: \"{entered_name}\"")
                user_confirmation = input("Czy wprowadzona nazwa jest poprawna? {T/N}: ")

                if user_confirmation.upper() == "T":
                    chosen_excel_file_name = f"{entered_name}.xlsx"
                    break
                else:
                    print("\nWystąpił błąd! Spróbuj ponownie!")
                    continue
            elif user_decision.upper() == "N":
                break
            else:
                print("\nWystąpił błąd! Spróbuj ponownie!")

        clear_console()

        return chosen_excel_file_name

    def get_excel_input_file(file_name):
        print("Wyszukuję wejściowego pliku arkusza kalkulacyjnego...")

        excel_file_name = None
        correct_file_found = False
        files_in_data_dir = os.listdir(DATA_PATH)

        for file in files_in_data_dir:
            if f"{file_name}".lower() in file.lower() and len(f"{file_name}") == len(
                    file):
                correct_file_found = True
                excel_file_name = file_name

                print("\nZnaleziono poprawny plik arkusza kalkulacyjnego...")

                break
            else:
                pass

        if not correct_file_found:
            print("\nNie znaleziono poprawnego pliku arkusza kalkulacyjnego! Umieść plik w folderze wejściowym!")
            print("Pamiętaj, że właściwy wejściowy plik arkusza kalkulacyjnego MUSI mieć rozszerzenie: \".xlsx\"!")
            input("Wciśnij \"Enter\" aby zamknąć program: ")
            leave()
        else:
            pass

        return excel_file_name

    correct_excel_file_name = verify_default_excel_file_name_correctness()
    excel_file = get_excel_input_file(correct_excel_file_name)

    excl = excel_file

    clear_console()

    return excl


def verify_output_path():
    print("Weryfikacja stanu folderu wyjściowego...")

    is_output_path = bool(os.path.isdir(OUTPUT_PATH))

    if not is_output_path:
        print("Nie odnaleziono folderu wyjściowego!")
        print("Tworzenie folderu wyjściowego...")

        os.makedirs(OUTPUT_PATH)
        is_output_dir_empty = True

        print("Utworzono folder wyjściowy...")

    else:
        is_output_dir_not_empty = bool(len(os.listdir(OUTPUT_PATH)))

        if not is_output_dir_not_empty:
            is_output_dir_empty = True
        else:
            is_output_dir_empty = False

    if not is_output_dir_empty:
        print("Folder wyjścia nie jest pusty! Program usunie wszystkie pliki w folderze wyjściowym.")
        _ask_user_if_one_wants_to_continue()

        shutil.rmtree(OUTPUT_PATH, ignore_errors=True)
        try:
            os.makedirs(OUTPUT_PATH)
        except FileExistsError:
            clear_console()
            print("Wystąpił błąd!\n\nZamknij wszystkie otwarte pliki z folderu wyjściowego!\n")
            input("Wciśnij \"Enter\" aby zamknąć program: ")
            leave()

        print("Usunięto pliki z folderu wyjściowego...")

    else:
        pass

    print("Pomyślnie zweryfikowano istnienie folderu wyjściowego...")

    clear_console()


def leave():
    for seconds in range(10):
        clear_console()
        print(f"Program zakończył działanie. Zamknięcie programu nastąpi za {10 - seconds} sekund...")
        time.sleep(1)

    sys.exit(0)


def clear_console():
    os.system("cls")


if __name__ == '__main__':
    inform_user_about_overwriting()

    excel = verify_environment()

    workbook = openpyxl.load_workbook(f"{DATA_PATH}/{excel}")
    worksheet = workbook[workbook.sheetnames[1]]

    process_data()

    format_output_file()

    save_excel_output_file()

    leave()
